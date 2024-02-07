from django.core.files.base import ContentFile
from django.shortcuts import get_object_or_404

from rest_framework import generics, status
from rest_framework.response import Response

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, Image
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


from core.models import SampleBoard, Marker
from core.serializers import SampleBoardSerializer
from .models import InstructionsPDF, Sticker
from .serializers import InstructionsPDFSerializer, StickerSerializer

import io

class InstructionsPDFListCreateAPIView(generics.ListCreateAPIView):
    queryset = InstructionsPDF.objects.all()
    serializer_class = InstructionsPDFSerializer

class InstructionsPDFDetailAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = InstructionsPDF.objects.all()
    serializer_class = InstructionsPDFSerializer

class StickerListCreateAPIView(generics.ListCreateAPIView):
    queryset = Sticker.objects.all()
    serializer_class = StickerSerializer

class StickerDetailAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Sticker.objects.all()
    serializer_class = StickerSerializer


class StickerListBySampleBoard(generics.ListAPIView):
    serializer_class = StickerSerializer

    def get_queryset(self):
        sample_board_uuid = self.kwargs['uuid']
        # Filter stickers based on the provided sample board id
        return Sticker.objects.filter(sample_board__uuid=sample_board_uuid)

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)

        # Get the sample board data
        sample_board_uuid = self.kwargs['uuid']
        sample_board = SampleBoard.objects.get(uuid=sample_board_uuid)
        sample_board_serializer = SampleBoardSerializer(sample_board)

        # Create a combined response with stickers and sample board data
        response_data = {
            
            'sample_board': sample_board_serializer.data,
            'stickers': serializer.data
        }

        return Response(response_data)

class InstructionsPDFGenerateAPIView(generics.CreateAPIView):
    serializer_class = InstructionsPDFSerializer

    def create(self, request, *args, **kwargs):
        excel_file = request.FILES.get('pdf_file')
        marker_id = request.data.get('marker')

        if not excel_file or not marker_id:
            return Response({'error': 'Excel file and marker_id are required.'}, status=status.HTTP_400_BAD_REQUEST)

        marker_instance = get_object_or_404(Marker, id=marker_id)

        workbook = load_workbook(excel_file)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            pdf_filename = f"{sheet_name}.pdf"

            # Convert sheet data to text
            text_data = "\n".join([str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value])

            # Convert text to image
            image_bytes = self.convert_text_to_image(text_data)

            # Save image as PDF file
            instructions_pdf_instance = InstructionsPDF(marker=marker_instance)
            instructions_pdf_instance.pdf_file.save(pdf_filename, ContentFile(image_bytes), save=True)

        return Response({'message': 'PDFs generated successfully.'}, status=status.HTTP_201_CREATED)

    def convert_text_to_image(self, text):
        # Create a PIL image canvas
        image_width = 800
        image_height = 600
        background_color = "white"
        text_color = "black"
        font_size = 18
        font_path = "arial.ttf"  # Path to your desired font file
        font = ImageFont.truetype(font_path, font_size)
        image = Image.new("RGB", (image_width, image_height), background_color)
        draw = ImageDraw.Draw(image)

        # Define text wrapping function
        def wrap_text(text, font, max_width):
            lines = []
            line = ""
            for word in text.split():
                if draw.textsize(line + word, font=font)[0] <= max_width:
                    line += " " + word
                else:
                    lines.append(line)
                    line = word
            lines.append(line)
            return lines

        # Split text into lines
        lines = wrap_text(text, font, image_width - 50)

        # Draw text onto image
        y = 20
        for line in lines:
            draw.text((20, y), line, fill=text_color, font=font)
            y += font_size

        # Save image as bytes
        image_bytes = io.BytesIO()
        image.save(image_bytes, format='PNG')
        image_bytes.seek(0)

        return image_bytes.getvalue()

